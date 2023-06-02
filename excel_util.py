from openpyxl import Workbook, load_workbook


# Method to read from excel
def get_ips():
    try:
        wb = load_workbook("details.xlsx")
        sheet = wb.active
        row_count = sheet.max_row + 1
        ips = [sheet[f'A{i}'].value for i in range(2, row_count)]
        return ips
    except BaseException as e:
        return e
